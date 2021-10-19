import csv
from openpyxl import load_workbook
import sqlite3


class BaseDados:
    conexao = sqlite3.connect('baseDados.db')
    cursor = conexao.cursor()

    cursor.execute('CREATE TABLE IF NOT EXISTS clientes ('
                   'cod'
                   'item'
                   'qtd total'
                   'qtd1'
                   'val1'
                   'qtd2'
                   'val2'
                   'qtd3'
                   'val3'
                   'qtd4'
                   'val4')

    cursor.execute('INSERT INTO clientes (cod,item,qtd total, qtd1, val1, qtd2, val2, qtd3, val3, qtd4, val4')





def salvarItem(temp, arquivo):
    with open(arquivo, 'a', newline='') as arquivo:
        escreve = csv.writer(
            arquivo,
            delimiter=',',  # Delimitador (Vírgula no caso)
            quotechar='"',  # Caractere de citação (No caso deixa os valores entre aspas)
            quoting=csv.QUOTE_ALL
        )
        try:
            escreve.writerow([temp[0], temp[1], temp[2], temp[3], temp[4]])
        except:
            escreve.writerow([temp[0], temp[1], temp[2]])

def getItens():

    wb = load_workbook(filename='tabela de preços 2021.xlsx', data_only=True)
    ws = wb.active
    row_count = int(ws.max_row)
    for c in range(1, row_count + 1):
        numA = "A" + str(c)
        numB = "B" + str(c)
        numC = "C" + str(c)
        cellA = ws[numA].value
        cellB = ws[numB].value
        cellC = ws[numC].value
        item = []
        if isinstance(cellA, int):
            item.append(cellA)
            item.append(cellB)
            item.append(str(cellC).replace('.', ','))
            salvarItem(item, 'itens.csv')
    for c in range(1, row_count + 1):
        numJ = "J" + str(c)
        numK = "K" + str(c)
        numL = "L" + str(c)
        cellJ = ws[numJ].value
        cellK = ws[numK].value
        cellL = ws[numL].value
        item = []
        if isinstance(cellJ, int):
            item.append(cellJ)
            item.append(cellK)
            item.append(str(cellL).replace('.', ','))
            salvarItem(item)

# getItens()